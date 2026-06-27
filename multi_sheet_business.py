import os
import duckdb
import uuid
import copy
from typing import Callable, Optional
from openpyxl import load_workbook, Workbook
from excel_tools import ExcelFileTools, DuckExcelEngine, get_high_fidelity_engine

class MultiSheetBusiness:
    """
    多表处理业务层：多文件管理、批量校验加载、多表组合运算
    与 WpsExcelBusiness 平级，纯业务逻辑，无UI依赖
    """
    def __init__(self, log_callback: Callable[[str], None]):
        self.log = log_callback
        self.file_paths: list[str] = []          # 已选文件绝对路径列表
        self.file_meta: dict[str, dict] = {}     # 文件元数据缓存 {path: 元数据}
        self.export_first_row_as_header: bool = True  # 导出是否将首行作为表头，与单表命名统一
        self.FILE_LIMIT: int = 100               # 文件数量软限制

    # ========== 文件列表管理 ==========
    def add_files(self, new_paths: list[str]) -> tuple[int, int]:
        """
        批量添加文件，自动去重，超上限抛出异常
        返回 (新增数量, 当前总数)
        """
        added = 0
        for p in new_paths:
            abs_path = os.path.abspath(p)
            if abs_path not in self.file_paths:
                if len(self.file_paths) >= self.FILE_LIMIT:
                    raise RuntimeError(f"文件数量已达上限 {self.FILE_LIMIT} 个，无法继续添加")
                self.file_paths.append(abs_path)
                added += 1
        return added, len(self.file_paths)

    def remove_files(self, indices: list[int]) -> int:
        """
        按索引移除选中文件，倒序删除避免索引错乱
        返回剩余文件数量
        """
        for i in sorted(indices, reverse=True):
            if 0 <= i < len(self.file_paths):
                removed_path = self.file_paths.pop(i)
                self.file_meta.pop(removed_path, None)
        return len(self.file_paths)

    def clear_all(self) -> None:
        """清空所有文件及元数据缓存"""
        self.file_paths.clear()
        self.file_meta.clear()

    # ========== 批量加载校验 ==========
    def load_all_files(self) -> dict:
        """
        后台批量校验所有文件、逐Sheet检测格式兼容性
        返回统计结果字典
        """
        valid_count = 0
        invalid_count = 0
        total_sheets = 0
        total_files = len(self.file_paths)
        for idx, path in enumerate(self.file_paths, 1):
            try:
                # 复用Excel文件基础校验
                is_valid, msg = ExcelFileTools.check_excel_valid(path)
                if not is_valid:
                    self.file_meta[path] = {
                        "valid": False,
                        "error": msg,
                        "all_non_empty_sheets": [],
                        "data_compatible_sheets": [],
                        "valid_sheets": [],
                        "sheet_count": 0,
                        "complex_reasons": {}
                    }
                    invalid_count += 1
                    self.log(f"[{idx}/{total_files}] {os.path.basename(path)} 校验失败：{msg}")
                    continue

                # 获取非空Sheet全集
                all_non_empty = ExcelFileTools.get_non_empty_sheets(path)
                # 统一检测数据模式兼容性
                compat_info = ExcelFileTools.check_sheet_data_compatibility(path)
                data_compatible = []
                complex_reasons = {}
                for name in all_non_empty:
                    info = compat_info.get(name, {"compatible": False, "reason": "检测失败"})
                    if info["compatible"]:
                        data_compatible.append(name)
                    else:
                        complex_reasons[name] = info["reason"]

                self.file_meta[path] = {
                    "valid": True,
                    "all_sheets": ExcelFileTools.list_sheet_names(path),
                    "all_non_empty_sheets": all_non_empty,
                    "data_compatible_sheets": data_compatible,
                    "valid_sheets": data_compatible,  # 向后兼容
                    "sheet_count": len(data_compatible),
                    "merged_sheets": list(complex_reasons.keys()),
                    "complex_reasons": complex_reasons
                }
                valid_count += 1
                total_sheets += len(data_compatible)
                
                # 输出逐文件明细日志
                total = len(all_non_empty)
                ok = len(data_compatible)
                complex_cnt = len(complex_reasons)
                self.log(f"[{idx}/{total_files}] {os.path.basename(path)} 加载完成：共{total}个非空Sheet，数据模式可用{ok}个，{complex_cnt}个仅支持保格式模式")
                if complex_reasons:
                    for name, reason in complex_reasons.items():
                        self.log(f"    - {name}：{reason}")
            except Exception as e:
                self.file_meta[path] = {
                    "valid": False,
                    "error": str(e),
                    "all_non_empty_sheets": [],
                    "data_compatible_sheets": [],
                    "valid_sheets": [],
                    "sheet_count": 0,
                    "complex_reasons": {}
                }
                invalid_count += 1
                self.log(f"[{idx}/{total_files}] {os.path.basename(path)} 加载异常：{str(e)}")
        
        # 汇总统计
        self.log(f"全部文件加载完成：共{total_files}个文件，有效{valid_count}个，失败{invalid_count}个，合计有效Sheet {total_sheets} 个")
        return {
            "total": total_files,
            "valid": valid_count,
            "invalid": invalid_count,
            "total_sheets": total_sheets
        }

    # ========== 预览辅助 ==========
    def get_preview_target(self, file_index: int = 0) -> tuple[Optional[str], Optional[str]]:
        """
        获取指定索引文件的第一个有效Sheet，用于预览
        返回 (文件路径, Sheet名称)，无效则返回None
        """
        if not self.file_paths or file_index < 0 or file_index >= len(self.file_paths):
            return None, None
        path = self.file_paths[file_index]
        meta = self.file_meta.get(path)
        if not meta or not meta["valid"] or not meta["valid_sheets"]:
            return None, None
        return path, meta["valid_sheets"][0]

    # ========== 输出路径辅助 ==========
    def _get_output_dir(self) -> str:
        """默认输出目录：第一个文件所在目录"""
        if not self.file_paths:
            return os.getcwd()
        return os.path.dirname(os.path.abspath(self.file_paths[0]))

    # ========== 内部工具：Sheet名称筛选 ==========
    def _filter_sheet_names_inner(self, original_sheets: list, keyword: str, mode: str) -> list:
        """
        业务层Sheet名称筛选工具
        6种匹配模式，全程严格区分大小写
        空关键词、空列表直接返回原列表，异常兜底返回原列表
        """
        try:
            if not keyword or not original_sheets:
                return original_sheets.copy()
            result = []
            for sheet_name in original_sheets:
                name_str = str(sheet_name)
                if mode == "equal":
                    match = (name_str == keyword)
                elif mode == "not_equal":
                    match = (name_str != keyword)
                elif mode == "contains":
                    match = (keyword in name_str)
                elif mode == "not_contains":
                    match = (keyword not in name_str)
                elif mode == "start_with":
                    match = name_str.startswith(keyword)
                elif mode == "end_with":
                    match = name_str.endswith(keyword)
                else:
                    # 未知模式直接返回原列表，不做任何过滤
                    return original_sheets.copy()
                if match:
                    result.append(sheet_name)
            return result
        except Exception:
            # 任何异常都兜底返回原列表，绝不中断主流程
            return original_sheets.copy()

    # ========== 功能1：合并所有文件为多Sheet大表 ==========
    def merge_all_files(self, output_path: str, sheet_filter_keyword: str = "", sheet_filter_mode: str = "contains", high_fidelity: bool = False) -> tuple[int, int, int]:
        """
        合并所有已加载文件的全部有效Sheet，输出为单个多Sheet Excel
        Sheet命名规则：文件名》Sheet名
        返回 (成功文件数, 成功Sheet数, 跳过Sheet数)
        """
        if not self.file_paths:
            raise RuntimeError("没有可合并的文件")
        return self._merge_file_list_inner(self.file_paths, output_path, sheet_filter_keyword, sheet_filter_mode, high_fidelity)

    # ========== 功能2：合并选中文件为多Sheet大表 ==========
    def merge_selected_files(self, indices: list[int], output_path: str, sheet_filter_keyword: str = "", sheet_filter_mode: str = "contains", high_fidelity: bool = False) -> tuple[int, int, int]:
        """合并指定索引的文件，逻辑与全量合并完全一致"""
        if not indices:
            raise RuntimeError("请先选择要合并的文件")
        selected_paths = [
            self.file_paths[i] 
            for i in indices 
            if 0 <= i < len(self.file_paths)
        ]
        if not selected_paths:
            raise RuntimeError("选中的文件无效")
        return self._merge_file_list_inner(selected_paths, output_path, sheet_filter_keyword, sheet_filter_mode, high_fidelity)

    # ========== 内部核心合并逻辑 ==========
    def _merge_file_list_inner(self, file_list: list[str], output_path: str,
                          sheet_filter_keyword: str = "", sheet_filter_mode: str = "contains",
                          high_fidelity: bool = False) -> tuple[int, int, int]:
        """
        多表格合并核心逻辑，按模式自动选择Sheet范围
        返回 (成功文件数, 成功Sheet数, 跳过Sheet数)
        """
        import tempfile
        # 启动清理临时文件
        cleaned = ExcelFileTools.cleanup_old_temp_files()
        if cleaned > 0:
            self.log(f"已自动清理 {cleaned} 个历史临时文件")
        # 输出路径前置校验
        output_dir = os.path.dirname(output_path)
        if not ExcelFileTools.is_dir_writable(output_dir):
            raise RuntimeError(f"输出目录无写入权限，请检查路径：{output_dir}")
        if ExcelFileTools.is_file_locked(output_path):
            raise RuntimeError("目标文件已被打开，请关闭 Excel 后重试")

        success_files = 0
        success_sheets = 0
        total_files = len(file_list)
        temp_files = []
        used_names = set()
        skipped_sheets = []
        temp_dir = tempfile.gettempdir()
        task_id = uuid.uuid4().hex[:8]

        if high_fidelity:
            self.log("启用保格式模式，所有非空Sheet均可参与合并，完整保留原表格样式...")
            engine = get_high_fidelity_engine()
            target_wb = Workbook()
            target_wb.remove(target_wb.active)
            try:
                for file_idx, file_path in enumerate(file_list, 1):
                    file_base = os.path.splitext(os.path.basename(file_path))[0]
                    meta = self.file_meta.get(file_path, {})
                    # 格式模式：使用所有非空Sheet全集
                    all_sheets = meta.get("all_non_empty_sheets", [])
                    # 先做名称筛选
                    filtered_sheets = self._filter_sheet_names_inner(all_sheets, sheet_filter_keyword, sheet_filter_mode)
                    skip_filter_count = len(all_sheets) - len(filtered_sheets)
                    
                    if skip_filter_count > 0:
                        for s in all_sheets:
                            if s not in filtered_sheets:
                                skipped_sheets.append(f"{file_base}》{s}（名称筛选不匹配）")
                    
                    if not filtered_sheets:
                        tip = "筛选后无有效Sheet，跳过" if skip_filter_count > 0 else "无有效Sheet，跳过"
                        self.log(f"[{file_idx}/{total_files}] ⏭️  {os.path.basename(file_path)} {tip}")
                        continue
                    
                    self.log(f"[{file_idx}/{total_files}] ▶️  正在处理 {os.path.basename(file_path)}（{len(filtered_sheets)}个Sheet）")
                    if skip_filter_count > 0:
                        self.log(f"[{file_idx}/{total_files}] 名称筛选已跳过 {skip_filter_count} 个Sheet")
                    
                    sheet_count_in_file = len(filtered_sheets)
                    file_process_ok = False
                    src_wb = load_workbook(file_path, keep_vba=True, data_only=False)
                    
                    try:
                        for sheet_name in filtered_sheets:
                            try:
                                # Sheet存在性校验
                                if sheet_name not in src_wb.sheetnames:
                                    skipped_sheets.append(f"{file_base}》{sheet_name}（Sheet不存在）")
                                    self.log(f"   ❌ Sheet[{sheet_name}] 不存在，已跳过")
                                    continue
                                
                                # 生成安全Sheet名并去重
                                raw_name = file_base if sheet_count_in_file == 1 else f"{file_base}》{sheet_name}"
                                safe_name = ExcelFileTools.sanitize_sheet_name(raw_name)
                                final_name = safe_name
                                dup_idx = 2
                                while final_name in used_names:
                                    suffix = f"({dup_idx})"
                                    base_len = 31 - len(suffix)
                                    final_name = safe_name[:base_len] + suffix if len(safe_name) > base_len else safe_name + suffix
                                    dup_idx += 1
                                used_names.add(final_name)
                                
                                # 核心：直接调用引擎通用复制方法
                                src_ws = src_wb[sheet_name]
                                new_ws = target_wb.create_sheet(title=final_name)
                                engine.copy_worksheet(src_ws, new_ws)
                                
                                success_sheets += 1
                                file_process_ok = True
                            except Exception as e:
                                err_msg = str(e)
                                skipped_sheets.append(f"{file_base}》{sheet_name}（处理失败：{err_msg}）")
                                self.log(f"   ❌ Sheet[{sheet_name}] 处理失败：{err_msg}")
                                continue
                    finally:
                        src_wb.close()
                    
                    if file_process_ok:
                        success_files += 1
                        self.log(f"[{file_idx}/{total_files}] ✅ {os.path.basename(file_path)} 处理完成")

                # 打印跳过详情
                if skipped_sheets:
                    self.log(f"📋 本次共跳过 {len(skipped_sheets)} 个Sheet：")
                    for item in skipped_sheets:
                        self.log(f"   - {item}")

                if success_sheets == 0:
                    raise RuntimeError("所有文件均无有效数据Sheet，合并失败")

                self.log("🔧 正在组装最终文件...")
                target_wb.save(output_path)
            finally:
                target_wb.close()
        else:
            # 数据模式：仅使用兼容Sheet
            self.log("数据模式，仅合并标准二维表结构的Sheet，含复杂格式的Sheet自动跳过")
            conn = duckdb.connect(":memory:")
            DuckExcelEngine._ensure_excel_extension(conn)
            try:
                for file_idx, file_path in enumerate(file_list, 1):
                    file_base = os.path.splitext(os.path.basename(file_path))[0]
                    meta = self.file_meta.get(file_path, {})
                    # 数据模式：仅使用兼容Sheet子集
                    data_sheets = meta.get("data_compatible_sheets", meta.get("valid_sheets", []))
                    complex_reasons = meta.get("complex_reasons", {})
                    
                    # 先做名称筛选（在兼容范围内筛选）
                    filtered_sheets = self._filter_sheet_names_inner(data_sheets, sheet_filter_keyword, sheet_filter_mode)
                    skip_filter_count = len(data_sheets) - len(filtered_sheets)
                    
                    # 记录复杂格式跳过的Sheet
                    all_non_empty = meta.get("all_non_empty_sheets", [])
                    for s in all_non_empty:
                        if s not in data_sheets:
                            skipped_sheets.append(f"{file_base}》{s}（{complex_reasons.get(s, '复杂格式不支持数据模式')}）")
                    if skip_filter_count > 0:
                        for s in data_sheets:
                            if s not in filtered_sheets:
                                skipped_sheets.append(f"{file_base}》{s}（名称筛选不匹配）")
                    
                    if not filtered_sheets:
                        self.log(f"[{file_idx}/{total_files}] ⏭️  {os.path.basename(file_path)} 数据模式下无有效Sheet，跳过")
                        continue
                    
                    self.log(f"[{file_idx}/{total_files}] ▶️  正在处理 {os.path.basename(file_path)}（{len(filtered_sheets)}个有效Sheet）")
                    
                    sheet_count_in_file = len(filtered_sheets)
                    file_process_ok = False
                    for sheet_name in filtered_sheets:
                        view_name = f"__tmp_merge_{success_sheets}__"
                        export_view = f"__merge_export_{success_sheets}__"
                        try:
                            if sheet_count_in_file == 1:
                                raw_name = file_base
                            else:
                                raw_name = f"{file_base}》{sheet_name}"
                            safe_name = ExcelFileTools.sanitize_sheet_name(raw_name)
                            final_name = safe_name
                            dup_idx = 2
                            while final_name in used_names:
                                suffix = f"({dup_idx})"
                                base_len = 31 - len(suffix)
                                final_name = safe_name[:base_len] + suffix if len(safe_name) > base_len else safe_name + suffix
                                dup_idx += 1
                            used_names.add(final_name)
                            DuckExcelEngine.read_clean_view(
                                conn, file_path, sheet_name,
                                header=False,
                                view_name=view_name,
                                dedup=True
                            )
                            row_count = conn.execute(f"SELECT COUNT(*) FROM {view_name}").fetchone()[0]
                            if row_count <= 0:
                                skipped_sheets.append(f"{file_base}》{sheet_name}（无数据）")
                                continue
                            if self.export_first_row_as_header and row_count < 2:
                                skipped_sheets.append(f"{file_base}》{sheet_name}（仅表头无数据）")
                                continue
                            data_rel = conn.sql(f"SELECT * FROM {view_name}")
                            if self.export_first_row_as_header:
                                first_row = data_rel.limit(1).fetchone()
                                if first_row:
                                    old_cols = data_rel.columns
                                    new_names = []
                                    name_counter = {}
                                    for idx, val in enumerate(first_row):
                                        raw_col_name = str(val).strip() if val is not None else ""
                                        raw_col_name = raw_col_name if raw_col_name else f"列{idx+1}"
                                        if raw_col_name in name_counter:
                                            name_counter[raw_col_name] += 1
                                            raw_col_name = f"{raw_col_name}_{name_counter[raw_col_name]}"
                                        else:
                                            name_counter[raw_col_name] = 0
                                        new_names.append(raw_col_name)
                                    rename_parts = [f'"{old}" AS "{new}"' for old, new in zip(old_cols, new_names)]
                                    col_str = ", ".join(rename_parts)
                                    data_rel = conn.sql(f"SELECT {col_str} FROM data_rel OFFSET 1")
                            temp_file_path = os.path.join(temp_dir, f"dkss_merge_{task_id}_{success_sheets}.xlsx")
                            conn.register(export_view, data_rel)
                            conn.execute(f"""
                                COPY (SELECT * FROM {export_view})
                                TO '{temp_file_path.replace(chr(39), chr(39)*2)}'
                                (
                                    FORMAT XLSX,
                                    SHEET '{final_name.replace(chr(39), chr(39)*2)}',
                                    HEADER TRUE,
                                    OVERWRITE TRUE
                                )
                            """)
                            temp_files.append((temp_file_path, final_name))
                            success_sheets += 1
                            file_process_ok = True
                        except Exception as e:
                            skipped_sheets.append(f"{file_base}》{sheet_name}（处理失败：{str(e)}）")
                            continue
                        finally:
                            try:
                                conn.execute(f'DROP VIEW IF EXISTS "{view_name}"')
                            except Exception:
                                pass
                            try:
                                conn.execute(f'DROP VIEW IF EXISTS "{export_view}"')
                            except Exception:
                                pass
                    if file_process_ok:
                        success_files += 1
                        self.log(f"[{file_idx}/{total_files}] ✅ {os.path.basename(file_path)} 处理完成")
                if success_sheets == 0:
                    raise RuntimeError("所有文件均无符合数据模式的有效Sheet，合并失败")
                self.log("🔧 正在组装最终文件...")
                ExcelFileTools.merge_excel_to_multi_sheet(
                    file_paths=[p for p, _ in temp_files],
                    output_path=output_path,
                    sheet_names=[s for _, s in temp_files]
                )
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
                for temp_path, _ in temp_files:
                    try:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    except Exception:
                        pass

        # 统一完整性校验
        try:
            wb = load_workbook(output_path, read_only=True)
            actual_count = len(wb.sheetnames)
            wb.close()
            if actual_count == success_sheets:
                self.log("✅ 完整性校验通过，Sheet数量与预期一致")
            else:
                self.log(f"⚠️ 警告：预期 {success_sheets} 个Sheet，实际 {actual_count} 个，结果可能不完整")
        except Exception as e:
            self.log(f"完整性校验跳过：{str(e)}")
        if skipped_sheets:
            self.log(f"📋 本次共跳过 {len(skipped_sheets)} 个Sheet：")
            for item in skipped_sheets:
                self.log(f"   - {item}")
        self.log(f"🎉 合并完成：成功 {success_files} 个文件、{success_sheets} 个Sheet")
        return success_files, success_sheets, len(skipped_sheets)
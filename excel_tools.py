import os
import re
from typing import Optional, Union, List, Dict, Tuple
import duckdb
import sys
from abc import ABC, abstractmethod
from copy import copy
from dataclasses import dataclass, field

# ========== 常量配置 ==========
EXCEL_SUFFIXES = (".xlsx", ".xlsm")
DEFAULT_SHEET_NAME = "Sheet1"
TEMP_SUFFIX = ".tmp.xlsx"

class ExcelFileTools:
    """
    Excel 文件级工具：纯标准库实现，零第三方依赖
    全部基于 ZIP 字节流扫描，极致性能，支持百MB级巨型文件
    """
    @staticmethod
    def is_excel_file(file_path: str) -> bool:
        """后缀 + 文件魔数双重校验，判断是否为合法 xlsx 文件"""
        if not file_path.lower().endswith(EXCEL_SUFFIXES):
            return False
        if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
            return False
        # xlsx 本质是 zip，魔数为 PK
        try:
            with open(file_path, "rb") as f:
                header = f.read(4)
            return header[:2] == b"PK"
        except Exception:
            return False

    @staticmethod
    def check_excel_valid(file_path: str) -> tuple[bool, str]:
        """Excel 文件专用有效性校验（纯标准库，不依赖DuckDB扩展）"""
        if not os.path.exists(file_path):
            return False, "文件不存在"
        if os.path.getsize(file_path) == 0:
            return False, "文件为空（字节大小为0）"
        if not file_path.lower().endswith(EXCEL_SUFFIXES):
            return False, "文件格式不支持，仅支持 .xlsx / .xlsm"
        
        try:
            import zipfile
            with zipfile.ZipFile(file_path, "r") as zf:
                # 检查xlsx核心结构文件是否存在
                has_workbook = "xl/workbook.xml" in zf.namelist()
                has_sheets = any(name.startswith("xl/worksheets/sheet") for name in zf.namelist())
            if not has_workbook or not has_sheets:
                return False, "文件结构损坏，不是合法的Excel文件"
            return True, "文件校验通过"
        except Exception as e:
            return False, f"Excel 文件读取失败：{str(e)}"

    @staticmethod
    def list_sheet_names(file_path: str) -> list[str]:
        """
        获取所有工作表名称（纯字节提取，零XML解析）
        从 xl/workbook.xml 中提取 name 属性值
        """
        import zipfile
        import re
        
        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                workbook_data = zf.read("xl/workbook.xml")
            
            # 正则提取所有 sheet 的 name 属性
            pattern = rb'<sheet[^>]*?name="([^"]+)"'
            matches = re.findall(pattern, workbook_data)
            return [name.decode("utf-8") for name in matches]
        except Exception as e:
            raise RuntimeError(f"读取Sheet列表失败：{str(e)}")

    @staticmethod
    def get_non_empty_sheets(file_path: str) -> list[str]:
        """
        获取所有非空Sheet名称列表（字节流扫描 <row 标签）
        有至少一行数据即判定为非空
        """
        import zipfile
        
        sheet_names = ExcelFileTools.list_sheet_names(file_path)
        result = []
        
        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                for idx, name in enumerate(sheet_names, 1):
                    sheet_path = f"xl/worksheets/sheet{idx}.xml"
                    if sheet_path not in zf.namelist():
                        continue
                    
                    # 分块流式搜索，不加载整个文件到内存
                    has_row = False
                    keyword = b"<row"
                    kw_len = len(keyword)
                    prev_tail = b""
                    
                    with zf.open(sheet_path, "r") as f:
                        while True:
                            chunk = f.read(1024 * 1024)  # 每次读1MB
                            if not chunk:
                                break
                            
                            search_buf = prev_tail + chunk
                            if keyword in search_buf:
                                has_row = True
                                break
                            
                            # 保留尾部，处理跨块边界
                            if len(chunk) >= kw_len - 1:
                                prev_tail = chunk[-(kw_len - 1):]
                            else:
                                prev_tail = chunk
                    
                    if has_row:
                        result.append(name)
            
            return result
        except Exception as e:
            raise RuntimeError(f"检测Sheet失败：{str(e)}")

    @staticmethod
    def check_sheet_merged_cells(file_path: str, sheet_name: str = None) -> dict:
        """
        检测指定Sheet或全部Sheet是否包含合并单元格
        纯字节流扫描 <mergeCells 标签，极致性能，零内存开销
        返回：{sheet名: 是否有合并单元格(bool)}
        """
        import zipfile
        
        all_sheet_names = ExcelFileTools.list_sheet_names(file_path)
        result = {}
        
        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                target_indices = []
                if sheet_name:
                    # 找到指定Sheet的序号
                    if sheet_name in all_sheet_names:
                        target_indices.append(all_sheet_names.index(sheet_name) + 1)
                    else:
                        raise RuntimeError(f"Sheet不存在：{sheet_name}")
                else:
                    target_indices = list(range(1, len(all_sheet_names) + 1))
                
                for idx in target_indices:
                    sheet_path = f"xl/worksheets/sheet{idx}.xml"
                    current_name = all_sheet_names[idx - 1]
                    
                    if sheet_path not in zf.namelist():
                        result[current_name] = False
                        continue
                    
                    # 分块流式搜索，不加载整个文件到内存
                    has_merged = False
                    keyword = b"<mergeCells"
                    kw_len = len(keyword)
                    prev_tail = b""
                    
                    with zf.open(sheet_path, "r") as f:
                        while True:
                            chunk = f.read(2 * 1024 * 1024)  # 每次读2MB
                            if not chunk:
                                break
                            
                            search_buf = prev_tail + chunk
                            if keyword in search_buf:
                                has_merged = True
                                break
                            
                            # 保留尾部，处理跨块边界
                            if len(chunk) >= kw_len - 1:
                                prev_tail = chunk[-(kw_len - 1):]
                            else:
                                prev_tail = chunk
                    
                    result[current_name] = has_merged
            
            return result
        except Exception as e:
            raise RuntimeError(f"合并单元格检测失败：{str(e)}")
        
    @staticmethod
    def check_sheet_data_compatibility(file_path: str, sheet_name: str = None) -> dict:
        """
        全项目统一的数据模式兼容性判定入口
        判定标准：会导致二维表结构错乱、列对齐失效的格式，视为数据模式不兼容
        当前规则：存在合并单元格 → 不兼容
        返回结构：{sheet名: {"compatible": bool, "reason": str}}
        """
        merged_info = ExcelFileTools.check_sheet_merged_cells(file_path, sheet_name)
        result = {}
        for name, has_merged in merged_info.items():
            if has_merged:
                result[name] = {
                    "compatible": False,
                    "reason": "存在合并单元格，列结构不标准，仅支持保格式模式"
                }
            else:
                result[name] = {
                    "compatible": True,
                    "reason": ""
                }
        return result

    @staticmethod
    def get_sheet_column_names(
        file_path: str, 
        sheet_name: str, 
        header: bool = True
    ) -> list[str]:
        """获取指定Sheet的列名列表（DuckDB读取，保证和后续处理列名一致）"""
        conn = duckdb.connect(":memory:")
        try:
            DuckExcelEngine._ensure_excel_extension(conn)
            sheet_param = f", sheet = '{sheet_name}'"
            header_param = f", header = {str(header).lower()}"
            rel = conn.sql(f"""
                SELECT *
                FROM read_xlsx('{file_path.replace(chr(39), chr(39)*2)}' {sheet_param} {header_param}, all_varchar = true)
                LIMIT 0
            """)
            return [col[0] for col in rel.description]
        finally:
            conn.close()
    
    @staticmethod
    def _num_to_col_letter(n: int) -> str:
        """数字转Excel风格列字母（0→a，1→b，26→aa）"""
        result = ""
        n += 1
        while n > 0:
            n -= 1
            result = chr(ord('a') + n % 26) + result
            n //= 26
        return result

    @staticmethod
    def generate_auto_column_names(sheet_index: int, col_count: int) -> list[str]:
        """
        生成无表头模式的列名，与UI下拉框完全一致
        sheet_index从0开始：第0个sheet a1,a2...，第1个b1,b2...
        """
        prefix = ExcelFileTools._num_to_col_letter(sheet_index)
        return [f"{prefix}{i+1}" for i in range(col_count)]
    
    @staticmethod
    def merge_excel_to_multi_sheet(file_paths: list[str], output_path: str, sheet_names: list[str] = None):
        """
        合并多个单Sheet Excel文件为一个多Sheet文件
        read_only 读源文件 + write_only 写目标文件，流式处理，内存占用极低
        :param file_paths: 单Sheet文件路径列表
        :param output_path: 输出文件路径
        :param sheet_names: 自定义Sheet名列表，为空则使用原Sheet名
        """
        from openpyxl import load_workbook, Workbook
        if not file_paths:
            raise ValueError("待合并文件列表不能为空")
        if sheet_names is None:
            sheet_names = [None] * len(file_paths)
        if len(file_paths) != len(sheet_names):
            raise ValueError("文件数量与Sheet名称数量不匹配")
        # 只写模式创建目标工作簿，本身无默认Sheet，无需手动删除
        wb_out = Workbook(write_only=True)
        try:
            for idx, file_path in enumerate(file_paths):
                # 只读模式打开源文件，仅读数据，不解析样式
                wb_in = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    ws_in = wb_in.active
                    target_sheet_name = sheet_names[idx] or ws_in.title
                    # 再次兜底Sheet名合规
                    target_sheet_name = ExcelFileTools.sanitize_sheet_name(target_sheet_name, f"Sheet{idx+1}")
                    ws_out = wb_out.create_sheet(title=target_sheet_name)
                    # 流式逐行写入，全程低内存
                    for row in ws_in.iter_rows(values_only=True):
                        ws_out.append(row)
                finally:
                    wb_in.close()
            wb_out.save(output_path)
        finally:
            wb_out.close()

    # === 全量合规 Sheet 命名（严格对齐微软官方规范）===
    @staticmethod
    def sanitize_sheet_name(name: str, default: str = "Sheet") -> str:
        r"""
        按 Excel 官方规则清洗 Sheet 名称
        规则：禁止 \ / ? * [ ] : 7个字符 | 最大31字符 | 不能以单引号首尾 | 不能为空白/历史记录
        """
        if not name or not str(name).strip():
            return default
        
        name = str(name).strip()
        # 替换所有非法字符为下划线
        name = re.sub(r'[\\/:*?\[\]]', '_', name)
        # 去除首尾单引号
        name = name.strip("'")
        # 过滤保留字「历史记录」
        if name == "历史记录" or name.lower() == "history":
            name = f"_{name}"
        # 长度截断至 31 字符
        if len(name) > 31:
            name = name[:31]
        # 最终兜底：处理完为空的情况
        if not name:
            return default
        return name

    # === 跨平台文件占用检测（Mac/Windows 通用，无第三方依赖）===
    @staticmethod
    def is_file_locked(file_path: str) -> bool:
        """检测文件是否被其他程序占用/无法写入"""
        if not os.path.exists(file_path):
            return False
        try:
            # 以追加独占模式尝试打开，成功则未被占用
            with open(file_path, 'a'):
                pass
            return False
        except (IOError, PermissionError):
            return True

    # === 输出目录可写性校验 ===
    @staticmethod
    def is_dir_writable(dir_path: str) -> bool:
        """检测目录是否存在且具备写入权限"""
        if not os.path.isdir(dir_path):
            try:
                os.makedirs(dir_path, exist_ok=True)
            except Exception:
                return False
        # 尝试创建临时文件验证写入权限
        test_file = os.path.join(dir_path, f".write_test_{os.getpid()}.tmp")
        try:
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
            return True
        except Exception:
            return False
        
    # === 体验优化：临时文件残留兜底清理 ===
    @staticmethod
    def cleanup_old_temp_files(prefix: str = "dkss_merge_", expire_hours: int = 24) -> int:
        """
        清理系统临时目录中过期的本程序临时文件
        :param prefix: 临时文件名统一前缀，精准匹配不误删
        :param expire_hours: 过期时长，默认24小时
        :return: 清理的文件数量
        """
        import tempfile
        import time
        temp_dir = tempfile.gettempdir()
        expire_timestamp = time.time() - expire_hours * 3600
        cleaned_count = 0
        try:
            for filename in os.listdir(temp_dir):
                # 只清理本模块生成的xlsx临时文件
                if not filename.startswith(prefix) or not filename.endswith(".xlsx"):
                    continue
                file_path = os.path.join(temp_dir, filename)
                try:
                    # 仅清理过期文件，跳过正在使用的文件
                    if os.path.getmtime(file_path) < expire_timestamp:
                        os.remove(file_path)
                        cleaned_count += 1
                except Exception:
                    continue
        except Exception:
            # 清理失败不影响主流程
            pass
        return cleaned_count
    
class DuckExcelEngine:
    """DuckDB 原生 Excel 读写引擎：读取、清洗、导出（纯数据处理）"""
    # ========== 扩展安装（强化错误提示版） ==========
    @staticmethod
    def _ensure_excel_extension(conn: duckdb.DuckDBPyConnection):
        """确保当前连接已加载 excel 扩展，失败直接抛出明确错误，绝不静默继续"""
        # 先尝试直接加载（已经安装过的情况）
        try:
            conn.execute("LOAD excel")
            return
        except Exception:
            pass
        # 加载失败，尝试安装
        try:
            conn.execute("INSTALL excel")
            conn.execute("LOAD excel")
            return
        except Exception as e:
            install_err = str(e)
            if "Failed to download" in install_err or "network" in install_err.lower():
                raise RuntimeError(
                    "DuckDB Excel 扩展安装失败：网络无法访问扩展仓库。\n"
                    "请检查网络连接，或手动执行 pip install --upgrade duckdb 升级到最新版本"
                )
            raise RuntimeError(f"DuckDB Excel 扩展加载失败：{install_err}")

    @staticmethod
    def peek_first_row(
        file: str,
        sheet_name: Optional[str] = None,
        header: bool = True
    ) -> list[str]:
        """轻量读取第一行数据，用于列预览"""
        conn = duckdb.connect(":memory:")
        try:
            DuckExcelEngine._ensure_excel_extension(conn)
            sheet_param = f", sheet = '{sheet_name}'" if sheet_name else ""
            header_param = f", header = {str(header).lower()}"
            
            rel = conn.sql(f"""
                SELECT *
                FROM read_xlsx('{file.replace(chr(39), chr(39)*2)}' {sheet_param} {header_param}, all_varchar = true)
                LIMIT 1
            """)
            rows = rel.fetchall()
            if not rows:
                return []
            return [str(v) if v is not None else '' for v in rows[0]]
        finally:
            conn.close()

    # ========== 读取清洗视图 ==========
    @staticmethod
    def read_clean_view(
        conn: duckdb.DuckDBPyConnection,
        file: str,
        sheet_name: Optional[str],
        header: bool,
        view_name: str,
        dedup: bool = True,
        custom_column_names: Optional[list[str]] = None
    ) -> list[str]:
        """
        读取 Excel 并生成清洗后的视图
        清洗规则：去首尾空格、空值转空串、过滤全空行、保序去重
        返回：列名列表
        """
        DuckExcelEngine._ensure_excel_extension(conn)
        sheet_param = f", sheet = '{sheet_name}'" if sheet_name else ""
        header_param = f", header = {str(header).lower()}"
        src_alias = f"{view_name}_src"
        # 读取原始数据，全部按字符串处理，避免手机号科学计数法
        rel = conn.sql(f"""
            SELECT *
            FROM read_xlsx('{file.replace(chr(39), chr(39)*2)}' {sheet_param} {header_param}, all_varchar = true)
        """)
        cols = [c[0] for c in rel.description]
        # 自定义列名重命名（无表头模式自动分配列名用）
        if custom_column_names and len(custom_column_names) == len(cols):
            rename_parts = [f'"{old}" AS "{new}"' for old, new in zip(cols, custom_column_names)]
            rel = rel.select(", ".join(rename_parts))
            cols = custom_column_names.copy()
        # 基础清洗：TRIM 去空格 + 空值转空串
        trim_parts = [
            f"COALESCE(TRIM(\"{c}\"::VARCHAR), '') AS \"{c}\""
            for c in cols
        ]
        rel = rel.select(", ".join(trim_parts))
        # 过滤全空行
        filter_parts = [f"\"{c}\" != ''" for c in cols]
        rel = rel.filter(" OR ".join(filter_parts))
        col_str = ", ".join([f'"{c}"' for c in cols])
        rel.create_view(src_alias)
        if dedup:
            # 整行全字段去重，保序
            sql_distinct = f"""
                SELECT {col_str}
                FROM (
                    SELECT *, ROW_NUMBER() OVER () AS __orig_row__
                    FROM {src_alias}
                )
                QUALIFY ROW_NUMBER() OVER (PARTITION BY {col_str} ORDER BY __orig_row__) = 1
                ORDER BY __orig_row__
            """
            rel_clean = conn.sql(sql_distinct)
        else:
            # 不去重，仅增加行号列
            sql = f"""
                SELECT {col_str}, ROW_NUMBER() OVER () AS __orig_row__
                FROM {src_alias}
            """
            rel_clean = conn.sql(sql)
        rel_clean.create_view(view_name)
        return cols

    # ========== 公开读取方法：多列 ==========
    @staticmethod
    def read_multi_column(
        conn: duckdb.DuckDBPyConnection,
        file: str,
        sheet_name: Optional[str] = None,
        header: bool = True
    ):
        """读取多列 Excel，返回原生 DuckDB Relation"""
        cols = DuckExcelEngine.read_clean_view(
            conn, file, sheet_name, header, "__excel_src__"
        )
        col_str = ", ".join([f'"{c}"' for c in cols])
        return conn.sql(f"SELECT {col_str} FROM __excel_src__")

    # ========== 公开读取方法：单列 ==========
    @staticmethod
    def read_single_column(
        conn: duckdb.DuckDBPyConnection,
        file: str,
        sheet_name: Optional[str] = None,
        header: bool = True,
        col_index: int = 0
    ):
        """读取指定列为单列数据集"""
        cols = DuckExcelEngine.read_clean_view(
            conn, file, sheet_name, header, "__excel_src__"
        )
        if col_index >= len(cols):
            raise RuntimeError(f"列索引超出范围，文件共 {len(cols)} 列")
        target_col = cols[col_index]
        sql = f"""
            SELECT "{target_col}" AS col0
            FROM __excel_src__
            WHERE "{target_col}" != ''
            QUALIFY ROW_NUMBER() OVER (PARTITION BY "{target_col}" ORDER BY __orig_row__) = 1
            ORDER BY __orig_row__
        """
        return conn.sql(sql)

    # ========== 按列名筛选读取 + 通用行过滤 ==========
    @staticmethod
    def read_filtered_columns(
        conn: duckdb.DuckDBPyConnection,
        file: str,
        sheet_name: str,
        target_columns: list[str],
        where_clause: str = "1=1",
        header: bool = True
    ) -> tuple[list[str], duckdb.DuckDBPyRelation]:
        """
        读取Sheet中指定的列，按WHERE条件筛选行
        保留原始行号用于后续多Sheet对齐拼接
        """
        DuckExcelEngine._ensure_excel_extension(conn)
        
        sheet_param = f", sheet = '{sheet_name}'"
        header_param = f", header = {str(header).lower()}"
        
        col_str = ", ".join(target_columns)
        # 子查询先生成原始行号，外层再做筛选，保证__rid__是原始行号
        base_sql = f"""
            SELECT *
            FROM (
                SELECT 
                    ROW_NUMBER() OVER () AS __rid__,
                    {col_str}
                FROM read_xlsx('{file.replace(chr(39), chr(39)*2)}' {sheet_param} {header_param}, all_varchar = true)
            ) t
            WHERE {where_clause}
        """
        
        rel = conn.sql(base_sql)
        return target_columns, rel

    # ========== 多Sheet同列数拼接 ==========
    @staticmethod
    def concat_same_column_sheets(
        conn: duckdb.DuckDBPyConnection,
        file: str,
        sheet_list: list[str],
        header: bool = True
    ) -> tuple[duckdb.DuckDBPyRelation, list[str], list[str]]:
        """
        按列名匹配拼接：提取所有Sheet共有的列首尾拼接
        不要求列数相同，仅拼接列名完全一致的共有列
        返回：(拼接后的Relation, 成功拼接的Sheet列表, 跳过的Sheet列表)
        """
        DuckExcelEngine._ensure_excel_extension(conn)
        if not sheet_list:
            raise ValueError("待拼接Sheet列表为空")
        
        # 1. 收集所有Sheet的列名，读取失败的自动跳过
        sheet_columns = {}
        skipped_sheets = []
        
        for sheet_name in sheet_list:
            try:
                cols = ExcelFileTools.get_sheet_column_names(file, sheet_name, header)
                sheet_columns[sheet_name] = cols
            except Exception:
                skipped_sheets.append(sheet_name)
        
        success_sheets = list(sheet_columns.keys())
        if len(success_sheets) < 2:
            raise RuntimeError("有效Sheet不足2个，无法拼接")
        
        # 2. 计算所有Sheet的共有列
        common_cols = set(sheet_columns[success_sheets[0]])
        for sheet_name in success_sheets[1:]:
            common_cols.intersection_update(sheet_columns[sheet_name])
        
        if not common_cols:
            raise RuntimeError("所选Sheet无共同列名，无法按列拼接")
        
        # 3. 按第一个Sheet的列顺序排列共有列，保持自然顺序
        base_cols = sheet_columns[success_sheets[0]]
        ordered_common = [col for col in base_cols if col in common_cols]
        col_str = ", ".join([f'"{c}"' for c in ordered_common])
        
        # 4. 构造每个Sheet的查询，只取共有列
        sql_parts = []
        header_param = f", header = {str(header).lower()}"
        
        for sheet_name in success_sheets:
            sheet_param = f", sheet = '{sheet_name}'"
            sql_parts.append(f"""
                SELECT {col_str}
                FROM read_xlsx('{file.replace(chr(39), chr(39)*2)}' {sheet_param} {header_param}, all_varchar = true)
            """)
        
        # 5. UNION ALL 拼接 + 统一清洗（去空格、过滤全空行、整行去重保序）
        union_sql = " UNION ALL ".join(sql_parts)
        clean_sql = f"""
            WITH raw_data AS ({union_sql})
            SELECT {col_str}
            FROM (
                SELECT 
                    {", ".join([f'COALESCE(TRIM("{c}"::VARCHAR), \'\') AS "{c}"' for c in ordered_common])},
                    ROW_NUMBER() OVER () AS __rid__
                FROM raw_data
            ) t
            WHERE {" OR ".join([f'"{c}" != \'\'' for c in ordered_common])}
            QUALIFY ROW_NUMBER() OVER (PARTITION BY {col_str} ORDER BY __rid__) = 1
            ORDER BY __rid__
        """
        return conn.sql(clean_sql), success_sheets, skipped_sheets

    # ========== 安全导出 Excel ==========
    @staticmethod
    def safe_to_excel(
        conn: duckdb.DuckDBPyConnection,
        rel: duckdb.DuckDBPyRelation,
        file_path: str,
        sheet_name: str = "Sheet1",
        header: bool = True
    ):
        """
        安全导出到 Excel：临时文件 + 原子替换
        中断/崩溃不会损坏原文件
        """
        DuckExcelEngine._ensure_excel_extension(conn)
        tmp_path = file_path + TEMP_SUFFIX
        header_param = str(header).lower()
        try:
            # 注册临时视图
            view_name = f"__export_{abs(hash(file_path))}__"
            conn.register(view_name, rel)
            
            conn.execute(f"""
                COPY (SELECT * FROM {view_name})
                TO '{tmp_path.replace(chr(39), chr(39)*2)}'
                (FORMAT XLSX, SHEET '{sheet_name}', HEADER {header_param})
            """)
            
            # 原子替换
            if not os.path.exists(tmp_path):
                raise RuntimeError("临时文件生成失败，导出终止")
            os.replace(tmp_path, file_path)
        except Exception as e:
            # 失败清理临时文件
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            raise RuntimeError(f"Excel 导出失败：{str(e)}")
        
class FilterBuilder:
    """高级筛选条件构建器：结构化条件 → 标准SQL WHERE子句，统一转义与校验"""
    # 算子映射：显示文本 → (内部标识, 值数量, SQL模板类型)
    OP_MAP = {
        "等于": ("eq", 1, "compare"),
        "大于": ("gt", 1, "compare"),
        "小于": ("lt", 1, "compare"),
        "大于等于": ("gte", 1, "compare"),
        "小于等于": ("lte", 1, "compare"),
        "不等于": ("ne", 1, "compare"),
        "包含": ("contains", 1, "like"),
        "不包含": ("not_contains", 1, "like"),
        "开头是": ("startswith", 1, "like"),
        "结尾是": ("endswith", 1, "like"),
        "为空": ("is_empty", 0, "null"),
        "非空": ("not_empty", 0, "null"),
        "长度等于": ("len_eq", 1, "length"),
        "长度大于": ("len_gt", 1, "length"),
        "长度小于": ("len_lt", 1, "length"),
        "属于": ("in", 1, "set"),
        "不属于": ("not_in", 1, "set"),
        "介于": ("between", 2, "range"),
    }
    COMPARE_OP = {
        "eq": "=",
        "gt": ">",
        "lt": "<",
        "gte": ">=",
        "lte": "<=",
        "ne": "<>"
    }
    LENGTH_OP = {
        "len_eq": "=",
        "len_gt": ">",
        "len_lt": "<"
    }
    @staticmethod
    def _is_number(s: str) -> bool:
        """判断字符串是否为有效数字"""
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False

    @staticmethod
    def _escape_sql_value(value: str) -> str:
        """SQL单引号转义"""
        return str(value).replace("'", "''")

    @staticmethod
    def _escape_like_pattern(value: str) -> str:
        """LIKE通配符转义，避免用户输入%_造成意外匹配"""
        escaped = value.replace("\\", "\\\\")
        escaped = escaped.replace("%", "\\%")
        escaped = escaped.replace("_", "\\_")
        return escaped

    @staticmethod
    def _trim_col(col: str) -> str:
        """统一对列值去首尾空格，所有算子共用"""
        return f"TRIM({col}::VARCHAR)"

    @staticmethod
    def build_condition_sql(col_name: str, op_text: str, val1: str = "", val2: str = "") -> str:
        """单个条件转SQL片段，列名必须是真实列名"""
        if op_text not in FilterBuilder.OP_MAP:
            raise ValueError(f"不支持的筛选算子：{op_text}")
        op_key, val_count, op_type = FilterBuilder.OP_MAP[op_text]
        col = f'"{col_name}"'
        trimmed_col = FilterBuilder._trim_col(col)
        if op_type == "null":
            if op_key == "is_empty":
                return f"({trimmed_col} = '')"
            else:
                return f"({trimmed_col} <> '')"
        if op_type == "compare":
            v = val1.strip()
            symbol = FilterBuilder.COMPARE_OP[op_key]
            # 统一按数值比较，失败则自动回退字符串比较
            if FilterBuilder._is_number(v):
                num_val = float(v)
                return f"(TRY_CAST({trimmed_col} AS DOUBLE) {symbol} {num_val})"
            else:
                escaped_v = FilterBuilder._escape_sql_value(v)
                return f"({trimmed_col} {symbol} '{escaped_v}')"
        if op_type == "like":
            v = val1.strip()
            escaped_v = FilterBuilder._escape_sql_value(FilterBuilder._escape_like_pattern(v))
            if op_key == "contains":
                pattern = f"%{escaped_v}%"
                return f"({trimmed_col} LIKE '{pattern}' ESCAPE '\\')"
            elif op_key == "not_contains":
                pattern = f"%{escaped_v}%"
                return f"({trimmed_col} NOT LIKE '{pattern}' ESCAPE '\\')"
            elif op_key == "startswith":
                pattern = f"{escaped_v}%"
                return f"({trimmed_col} LIKE '{pattern}' ESCAPE '\\')"
            elif op_key == "endswith":
                pattern = f"%{escaped_v}"
                return f"({trimmed_col} LIKE '{pattern}' ESCAPE '\\')"
            else:
                pattern = f"%{escaped_v}%"
                return f"({trimmed_col} LIKE '{pattern}' ESCAPE '\\')"
        if op_type == "length":
            if not val1.strip().isdigit():
                raise ValueError("长度比较必须输入正整数")
            symbol = FilterBuilder.LENGTH_OP[op_key]
            len_val = int(val1.strip())
            return f"(LENGTH({trimmed_col}) {symbol} {len_val})"
        if op_type == "set":
            items = [x.strip() for x in val1.split(",") if x.strip()]
            if not items:
                raise ValueError("属于/不属于必须输入至少一个值，用英文逗号分隔")
            escaped_items = [f"'{FilterBuilder._escape_sql_value(x)}'" for x in items]
            items_str = ", ".join(escaped_items)
            if op_key == "in":
                return f"({trimmed_col} IN ({items_str}))"
            else:
                return f"({trimmed_col} NOT IN ({items_str}))"
        if op_type == "range":
            v1 = val1.strip()
            v2 = val2.strip()
            if FilterBuilder._is_number(v1) and FilterBuilder._is_number(v2):
                num1 = float(v1)
                num2 = float(v2)
                return f"(TRY_CAST({trimmed_col} AS DOUBLE) BETWEEN {num1} AND {num2})"
            else:
                ev1 = FilterBuilder._escape_sql_value(v1)
                ev2 = FilterBuilder._escape_sql_value(v2)
                return f"({trimmed_col} BETWEEN '{ev1}' AND '{ev2}')"
        return "1=1"

    @staticmethod
    def build_group_sql(conditions: list[dict]) -> str:
        """条件组转SQL：组内所有条件AND连接"""
        if not conditions:
            return "1=1"
        sql_parts = []
        for cond in conditions:
            sql_parts.append(FilterBuilder.build_condition_sql(
                cond["col"], cond["op"], cond.get("val", ""), cond.get("val2", "")
            ))
        return f"({' AND '.join(sql_parts)})"

    @staticmethod
    def build_groups_sql(groups: list[list[dict]]) -> str:
        """多条件组转SQL：组间OR连接"""
        if not groups:
            return "1=1"
        group_sqls = [FilterBuilder.build_group_sql(g) for g in groups if g]
        if not group_sqls:
            return "1=1"
        if len(group_sqls) == 1:
            return group_sqls[0]
        return " OR ".join(group_sqls)

# ========== 保格式引擎（openpyxl 统一实现，跨平台一致）==========
class BaseHighFidelityEngine(ABC):
    """保格式Excel引擎抽象基类，拆分/合并接口统一"""
    @abstractmethod
    def split_sheets(
        self,
        file_path: str,
        sheet_names: list[str],
        output_dir: str,
        name_template: str = "{file_base}-{sheet}.xlsx"
    ) -> int:
        """保格式拆分指定Sheet为独立文件，返回成功数量"""
        pass

    @abstractmethod
    def merge_to_multi_sheet(
        self,
        file_list: list[str],
        output_path: str,
        sheet_name_list: list[str] = None
    ) -> int:
        """保格式合并多文件为多Sheet单文件，返回成功Sheet数量"""
        pass

class OpenpyxlHighFidelityEngine(BaseHighFidelityEngine):
    """基于 openpyxl 的保格式引擎，全平台通用，统一维护一份复制逻辑"""

    @staticmethod
    def copy_worksheet(src_ws, dst_ws):
        """
        跨工作簿高保真完整复制工作表（最终可用版）
        修复条件格式范围提取，全格式完整还原，全链路异常兜底
        """
        from copy import copy
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection

        # 差异样式类导入（官方标准路径）
        DifferentialStyle = None
        try:
            from openpyxl.styles.differential import DifferentialStyle
        except ImportError:
            DifferentialStyle = None

        # ========== 1. Sheet级基础属性 ==========
        try:
            dst_ws.sheet_format = copy(src_ws.sheet_format)
            dst_ws.sheet_properties = copy(src_ws.sheet_properties)
            dst_ws.sheet_view = copy(src_ws.sheet_view)
            dst_ws.page_margins = copy(src_ws.page_margins)
            dst_ws.page_setup = copy(src_ws.page_setup)
            dst_ws.print_options = copy(src_ws.print_options)
            dst_ws.freeze_panes = src_ws.freeze_panes
            if hasattr(src_ws, 'print_title_rows'):
                dst_ws.print_title_rows = src_ws.print_title_rows
            if hasattr(src_ws, 'print_title_cols'):
                dst_ws.print_title_cols = src_ws.print_title_cols
            if hasattr(src_ws, 'print_area'):
                dst_ws.print_area = src_ws.print_area
        except Exception:
            pass

        # ========== 2. 单元格级：值、公式、全量样式 ==========
        for row in src_ws.iter_rows():
            for cell in row:
                try:
                    new_cell = dst_ws.cell(row=cell.row, column=cell.column)

                    # 值与公式
                    if hasattr(cell, 'array_formula') and cell.array_formula:
                        new_cell.array_formula = cell.array_formula
                    elif cell.data_type == 'f':
                        new_cell.value = cell.value
                    elif hasattr(cell, 'rich_text') and cell.rich_text:
                        try:
                            new_cell.rich_text = copy(cell.rich_text)
                        except Exception:
                            new_cell.value = cell.value
                    else:
                        new_cell.value = cell.value

                    # 数字格式
                    new_cell.number_format = cell.number_format

                    # 样式拆解重建
                    if cell.has_style:
                        # 字体
                        try:
                            sf = cell.font
                            new_cell.font = Font(
                                name=sf.name, size=sf.size, bold=sf.bold, italic=sf.italic,
                                underline=sf.underline, strike=sf.strike, color=copy(sf.color),
                                vertAlign=sf.vertAlign, outline=sf.outline, shadow=sf.shadow,
                                condense=sf.condense, extend=sf.extend, family=sf.family,
                                charset=sf.charset, scheme=sf.scheme
                            )
                        except Exception:
                            pass

                        # 填充
                        try:
                            sf = cell.fill
                            new_cell.fill = PatternFill(
                                fill_type=sf.fill_type,
                                start_color=copy(sf.start_color),
                                end_color=copy(sf.end_color)
                            )
                        except Exception:
                            pass

                        # 边框
                        try:
                            sb = cell.border
                            def _copy_side(side):
                                return Side(
                                    style=side.style, color=copy(side.color),
                                    border_style=side.border_style
                                )
                            new_cell.border = Border(
                                left=_copy_side(sb.left), right=_copy_side(sb.right),
                                top=_copy_side(sb.top), bottom=_copy_side(sb.bottom),
                                diagonal=_copy_side(sb.diagonal),
                                diagonal_direction=sb.diagonal_direction,
                                outline=sb.outline, vertical=sb.vertical, horizontal=sb.horizontal
                            )
                        except Exception:
                            pass

                        # 对齐
                        try:
                            sa = cell.alignment
                            new_cell.alignment = Alignment(
                                horizontal=sa.horizontal, vertical=sa.vertical,
                                text_rotation=sa.text_rotation, wrap_text=sa.wrap_text,
                                shrink_to_fit=sa.shrink_to_fit, indent=sa.indent,
                                relativeIndent=sa.relativeIndent,
                                justifyLastLine=sa.justifyLastLine,
                                readingOrder=sa.readingOrder
                            )
                        except Exception:
                            pass

                        # 单元格保护
                        try:
                            sp = cell.protection
                            new_cell.protection = Protection(
                                locked=sp.locked, hidden=sp.hidden
                            )
                        except Exception:
                            pass

                    # 超链接
                    if cell.hyperlink:
                        try:
                            new_cell.hyperlink = copy(cell.hyperlink)
                        except Exception:
                            pass

                    # 批注
                    if cell.comment:
                        try:
                            new_cell.comment = copy(cell.comment)
                        except Exception:
                            pass
                except Exception:
                    continue

        # ========== 3. 行高列宽 + 分级显示 ==========
        try:
            for row_idx, row_dim in src_ws.row_dimensions.items():
                dst_ws.row_dimensions[row_idx].height = row_dim.height
                dst_ws.row_dimensions[row_idx].hidden = row_dim.hidden
                dst_ws.row_dimensions[row_idx].outline_level = row_dim.outline_level
                dst_ws.row_dimensions[row_idx].collapsed = row_dim.collapsed
            for col_letter, col_dim in src_ws.column_dimensions.items():
                dst_ws.column_dimensions[col_letter].width = col_dim.width
                dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden
                dst_ws.column_dimensions[col_letter].outline_level = col_dim.outline_level
                dst_ws.column_dimensions[col_letter].collapsed = col_dim.collapsed
        except Exception:
            pass

        # ========== 4. 合并单元格 ==========
        try:
            for merged_range in src_ws.merged_cells.ranges:
                dst_ws.merge_cells(str(merged_range))
        except Exception:
            pass

        # ========== 5. 自动筛选 + 排序状态 ==========
        try:
            if src_ws.auto_filter.ref:
                dst_ws.auto_filter.ref = src_ws.auto_filter.ref
                dst_ws.auto_filter.filterColumn = copy(src_ws.auto_filter.filterColumn)
                dst_ws.auto_filter.sortState = copy(src_ws.auto_filter.sortState)
        except Exception:
            pass

        # ========== 6. 条件格式（最终修复版：正确提取纯范围）==========
        if DifferentialStyle is not None:
            try:
                from openpyxl.formatting.rule import (
                    CellIsRule, ColorScaleRule, DataBarRule, IconSetRule,
                    FormulaRule, Rule
                )
                cf_dict = src_ws.conditional_formatting._cf_rules
                for cf_obj, rule_list in cf_dict.items():
                    # 正确提取纯单元格范围字符串
                    try:
                        range_str = str(cf_obj.sqref).strip()
                    except Exception:
                        range_str = str(cf_obj).replace('<ConditionalFormatting ', '').replace('>', '').strip()
                    if not range_str:
                        continue

                    for rule in rule_list:
                        try:
                            # 重建差异样式
                            new_dxf = None
                            if rule.dxf:
                                dxf = rule.dxf
                                new_dxf = DifferentialStyle()
                                try:
                                    if dxf.font:
                                        f = dxf.font
                                        new_dxf.font = Font(
                                            name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                                            underline=f.underline, strike=f.strike, color=copy(f.color)
                                        )
                                except Exception:
                                    pass
                                try:
                                    if dxf.fill:
                                        fl = dxf.fill
                                        new_dxf.fill = PatternFill(
                                            fill_type=fl.fill_type,
                                            start_color=copy(fl.start_color),
                                            end_color=copy(fl.end_color)
                                        )
                                except Exception:
                                    pass
                                try:
                                    if dxf.border:
                                        b = dxf.border
                                        def _copy_side_dxf(side):
                                            return Side(style=side.style, color=copy(side.color))
                                        new_dxf.border = Border(
                                            left=_copy_side_dxf(b.left), right=_copy_side_dxf(b.right),
                                            top=_copy_side_dxf(b.top), bottom=_copy_side_dxf(b.bottom)
                                        )
                                except Exception:
                                    pass

                            rule_type = rule.type
                            stop_if = getattr(rule, 'stopIfTrue', False)

                            # 按类型重建规则
                            if rule_type == 'cellIs':
                                new_rule = CellIsRule(
                                    operator=rule.operator, formula=rule.formula,
                                    stopIfTrue=stop_if, dxf=new_dxf
                                )
                            elif rule_type == 'expression':
                                new_rule = FormulaRule(
                                    formula=rule.formula, stopIfTrue=stop_if, dxf=new_dxf
                                )
                            elif rule_type == 'colorScale':
                                cs = rule.colorScale
                                cfvos = cs.cfvo
                                colors = cs.color
                                start_type = cfvos[0].type
                                start_value = cfvos[0].val
                                start_color = colors[0].value if hasattr(colors[0], 'value') else str(colors[0])

                                if len(cfvos) >= 3:
                                    mid_type = cfvos[1].type
                                    mid_value = cfvos[1].val
                                    mid_color = colors[1].value if hasattr(colors[1], 'value') else str(colors[1])
                                    end_type = cfvos[2].type
                                    end_value = cfvos[2].val
                                    end_color = colors[2].value if hasattr(colors[2], 'value') else str(colors[2])
                                    new_rule = ColorScaleRule(
                                        start_type=start_type, start_value=start_value, start_color=start_color,
                                        mid_type=mid_type, mid_value=mid_value, mid_color=mid_color,
                                        end_type=end_type, end_value=end_value, end_color=end_color,
                                        stopIfTrue=stop_if
                                    )
                                else:
                                    end_type = cfvos[1].type
                                    end_value = cfvos[1].val
                                    end_color = colors[1].value if hasattr(colors[1], 'value') else str(colors[1])
                                    new_rule = ColorScaleRule(
                                        start_type=start_type, start_value=start_value, start_color=start_color,
                                        end_type=end_type, end_value=end_value, end_color=end_color,
                                        stopIfTrue=stop_if
                                    )
                            elif rule_type == 'dataBar':
                                db = rule.dataBar
                                cfvos = db.cfvo
                                bar_color = db.color.value if hasattr(db.color, 'value') else str(db.color)
                                new_rule = DataBarRule(
                                    start_type=cfvos[0].type,
                                    start_value=cfvos[0].val,
                                    end_type=cfvos[1].type,
                                    end_value=cfvos[1].val,
                                    color=bar_color,
                                    showValue=db.showValue,
                                    minLength=getattr(db, 'minLength', None),
                                    maxLength=getattr(db, 'maxLength', None),
                                    stopIfTrue=stop_if
                                )
                            elif rule_type == 'iconSet':
                                icons = rule.iconSet
                                values = [vo.val for vo in icons.cfvo]
                                new_rule = IconSetRule(
                                    iconSet=icons.iconSet,
                                    showValue=icons.showValue,
                                    percent=icons.percent,
                                    reverse=icons.reverse,
                                    values=values,
                                    stopIfTrue=stop_if
                                )
                            else:
                                # 通用规则全属性复制兜底
                                new_rule = Rule(type=rule_type, dxf=new_dxf, stopIfTrue=stop_if)
                                for attr in dir(rule):
                                    if attr.startswith('_') or attr in ('dxf', 'stopIfTrue', 'type'):
                                        continue
                                    try:
                                        setattr(new_rule, attr, copy(getattr(rule, attr)))
                                    except Exception:
                                        pass

                            dst_ws.conditional_formatting.add(range_str, new_rule)
                        except Exception:
                            # 单条规则失败跳过
                            continue
            except Exception:
                # 整个条件格式模块异常兜底
                pass

        # ========== 7. 数据验证 ==========
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            for dv in src_ws.data_validations.dataValidation:
                try:
                    new_dv = DataValidation(
                        type=dv.type, operator=dv.operator,
                        formula1=dv.formula1, formula2=dv.formula2,
                        allow_blank=dv.allow_blank, showDropDown=dv.showDropDown,
                        showInputMessage=dv.showInputMessage,
                        promptTitle=dv.promptTitle, prompt=dv.prompt,
                        showErrorMessage=dv.showErrorMessage,
                        errorTitle=dv.errorTitle, error=dv.error,
                        errorStyle=dv.errorStyle
                    )
                    for sqref in dv.sqref:
                        new_dv.add(str(sqref))
                    dst_ws.add_data_validation(new_dv)
                except Exception:
                    continue
        except Exception:
            pass

        # ========== 8. 迷你图 ==========
        try:
            from openpyxl.worksheet.sparkline import SparklineGroup, Sparkline
            for group in src_ws.sparkline_groups:
                try:
                    sparklines = []
                    for sp in group.sparklines:
                        sparklines.append(Sparkline(
                            location=sp.location, dataRange=sp.dataRange
                        ))
                    new_group = SparklineGroup(
                        sparklines=sparklines,
                        type=group.type,
                        displayEmptyCellsAs=group.displayEmptyCellsAs,
                        markers=group.markers, high=group.high, low=group.low,
                        first=group.first, last=group.last, negative=group.negative,
                        displayXAxis=group.displayXAxis, displayHidden=group.displayHidden,
                        minAxisType=group.minAxisType, maxAxisType=group.maxAxisType,
                        manualMin=group.manualMin, manualMax=group.manualMax,
                        rightToLeft=group.rightToLeft,
                        colorSeries=copy(group.colorSeries),
                        colorNegative=copy(group.colorNegative),
                        colorMarkers=copy(group.colorMarkers),
                        colorFirst=copy(group.colorFirst),
                        colorLast=copy(group.colorLast),
                        colorHigh=copy(group.colorHigh),
                        colorLow=copy(group.colorLow)
                    )
                    dst_ws.sparkline_groups.append(new_group)
                except Exception:
                    continue
        except Exception:
            pass

        # ========== 9. 工作表保护 ==========
        try:
            src_prot = src_ws.protection
            dst_ws.protection.sheet = src_prot.sheet
            dst_ws.protection.objects = src_prot.objects
            dst_ws.protection.scenarios = src_prot.scenarios
            dst_ws.protection.format_cells = src_prot.format_cells
            dst_ws.protection.format_columns = src_prot.format_columns
            dst_ws.protection.format_rows = src_prot.format_rows
            dst_ws.protection.insert_columns = src_prot.insert_columns
            dst_ws.protection.insert_rows = src_prot.insert_rows
            dst_ws.protection.insert_hyperlinks = src_prot.insert_hyperlinks
            dst_ws.protection.delete_columns = src_prot.delete_columns
            dst_ws.protection.delete_rows = src_prot.delete_rows
            dst_ws.protection.select_locked_cells = src_prot.select_locked_cells
            dst_ws.protection.sort = src_prot.sort
            dst_ws.protection.auto_filter = src_prot.auto_filter
            dst_ws.protection.pivot_tables = src_prot.pivot_tables
            dst_ws.protection.select_unlocked_cells = src_prot.select_unlocked_cells
            if hasattr(src_prot, 'password') and src_prot.password:
                dst_ws.protection.password = src_prot.password
        except Exception:
            pass

    def split_sheets(self, src_file: str, sheet_names: list[str], output_dir: str, name_template: str = None):
        """
        拆分指定Sheet为独立Excel文件，基于copy_worksheet全量高保真复制
        完整保留：单元格样式、公式、合并单元格、条件格式、数据验证、冻结窗格、页面设置、工作表保护
        """
        from openpyxl import load_workbook, Workbook
        import os

        src_wb = load_workbook(src_file, keep_vba=True, data_only=False)
        try:
            for sheet_name in sheet_names:
                if sheet_name not in src_wb.sheetnames:
                    continue
                src_ws = src_wb[sheet_name]

                # 新建目标工作簿，移除默认空白Sheet
                dst_wb = Workbook()
                dst_wb.remove(dst_wb.active)

                # 创建目标Sheet，复用全量高保真复制方法
                dst_ws = dst_wb.create_sheet(title=sheet_name)
                self.copy_worksheet(src_ws, dst_ws)

                # 生成输出路径
                if name_template:
                    out_filename = name_template
                else:
                    out_filename = f"{sheet_name}.xlsx"
                out_path = os.path.join(output_dir, out_filename)

                dst_wb.save(out_path)
                dst_wb.close()
        finally:
            src_wb.close()

    def merge_to_multi_sheet(self, file_list: list[str], output_path: str,
                             sheet_name_list: list[str] = None) -> int:
        from openpyxl import load_workbook, Workbook
        if not file_list:
            raise ValueError("待合并文件列表不能为空")
        if sheet_name_list is None:
            sheet_name_list = [None] * len(file_list)
        if len(file_list) != len(sheet_name_list):
            raise ValueError("文件数量与Sheet名称数量不匹配")
        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        success = 0
        try:
            for idx, file_path in enumerate(file_list):
                wb_in = load_workbook(file_path, data_only=False, keep_vba=True)
                try:
                    ws_in = wb_in.active
                    target_name = sheet_name_list[idx] or ws_in.title
                    target_name = ExcelFileTools.sanitize_sheet_name(target_name, f"Sheet{idx+1}")
                    ws_out = wb_out.create_sheet(title=target_name)
                    # 复用通用复制方法
                    self.copy_worksheet(ws_in, ws_out)
                    success += 1
                finally:
                    wb_in.close()
            wb_out.save(output_path)
        finally:
            wb_out.close()
        return success

def get_high_fidelity_engine() -> BaseHighFidelityEngine:
    """保格式模式统一使用 openpyxl 实现，全平台行为一致"""
    return OpenpyxlHighFidelityEngine()

class SheetCheckResult:
    sheet_name: str
    is_standard: bool
    # 强否定原因：直接导致数据模式不可用
    reject_reasons: List[str] = field(default_factory=list)
    # 弱提示原因：不强制禁用，仅预警风险
    warn_reasons: List[str] = field(default_factory=list)

class ExcelStandardChecker:
    """
    基于DuckDB数据引擎的能力边界，检测Excel Sheet是否为标准可运算二维表
    判定标准：DuckDB read_xlsx能否将Sheet正确读取为可SQL运算的规整二维表
    """

    # 扫描前N行做结构判定，避免大文件性能损耗
    SCAN_MAX_ROW = 200
    # 列数波动阈值：非空列数最大最小差超过该值，判定为不规则排版
    COL_DIFF_THRESHOLD = 3

    @classmethod
    def check_file(cls, file_path: str) -> Dict[str, SheetCheckResult]:
        """检测文件中所有非空Sheet，返回每个Sheet的检测结果"""
        result = {}
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row == 0 or ws.max_column == 0:
                    continue
                result[sheet_name] = cls._check_sheet(ws)
            wb.close()
        except Exception:
            # 读取失败时默认按非标准处理，由上层业务兜底
            pass
        return result

    @classmethod
    def _check_sheet(cls, ws) -> SheetCheckResult:
        sheet_name = ws.title
        result = SheetCheckResult(sheet_name=sheet_name, is_standard=True)
        has_merge = bool(ws.merged_cells.ranges)

        # 强否定规则1：存在合并单元格
        if has_merge:
            result.is_standard = False
            result.reject_reasons.append("存在合并单元格，破坏行列对应关系")

        # 扫描前SCAN_MAX_ROW行做结构检测
        scan_rows = min(ws.max_row, cls.SCAN_MAX_ROW)
        col_counts = []
        has_empty_row_gap = False
        last_non_empty_row = 0

        for row_idx in range(1, scan_rows + 1):
            row = ws[row_idx]
            non_empty_count = sum(1 for cell in row if cell.value is not None)
            if non_empty_count > 0:
                col_counts.append(non_empty_count)
                if row_idx - last_non_empty_row > 2 and last_non_empty_row > 0:
                    has_empty_row_gap = True
                last_non_empty_row = row_idx

        # 强否定规则2：行列结构严重不规则
        if col_counts:
            max_col = max(col_counts)
            min_col = min(col_counts)
            if max_col - min_col >= cls.COL_DIFF_THRESHOLD:
                result.is_standard = False
                result.reject_reasons.append("行列结构不规则，非标准平铺一维表")

        # 强否定规则3：多层表头/首行非有效列名
        if has_merge and scan_rows >= 5:
            # 前5行存在跨列合并且首行非空单元格极少，判定为多层表头
            first_row_non_empty = sum(1 for cell in ws[1] if cell.value is not None)
            if first_row_non_empty < max(col_counts) // 2:
                result.is_standard = False
                if "存在多层表头结构" not in result.reject_reasons:
                    result.reject_reasons.append("存在多层表头结构，首行非有效列名")

        # 弱提示规则
        if has_empty_row_gap:
            result.warn_reasons.append("数据区域存在连续空行，可能影响读取范围")
        if ws.max_row > scan_rows:
            result.warn_reasons.append(f"仅扫描前{cls.SCAN_MAX_ROW}行，后续行结构可能存在差异")

        return result

    @classmethod
    def get_file_aggregate_status(cls, check_result: Dict[str, SheetCheckResult]) -> Tuple[bool, bool]:
        """
        聚合文件级检测状态
        返回：(是否存在非标准Sheet, 是否所有Sheet均为非标准)
        """
        if not check_result:
            return True, True
        has_non_standard = any(not r.is_standard for r in check_result.values())
        all_non_standard = all(not r.is_standard for r in check_result.values())
        return has_non_standard, all_non_standard